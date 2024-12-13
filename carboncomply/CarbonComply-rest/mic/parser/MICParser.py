from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from exceptions.ParserException import ParserException
from utils.NormalizeText import normalize_list

mic_original_header = ['Period', 'ATF Number', 'Customs Clearance Office', 'Country of Dispatch', 'Document Number', 'Currency',
              'Invoice Number', 'Customs Clearance Date',  'Invoice Date', 'Supplier Number', 'Supplier Name',
              'Article Number', 'Article Description', 'Quantity', 'Own Mass', 'Customs Clearance Date',
              'Country of Origin', 'CN', 'Procedure Code', 'Requested Procedure', 'Previous Procedure', 'VAT Base',
              'VAT Amount', 'Currency VAT', 'Customs Duty Base', 'Customs Duty Amount', 'Currency Duty',
              'Position System ID', 'Material Number']

mic_header = ['period', 'atf_number', 'customs_clearance_office', 'country_of_dispatch', 'document_number',
                         'currency', 'invoice_number', 'customs_clearance_date', 'invoice_date', 'supplier_number',
                         'supplier_name', 'article_number', 'article_description', 'quantity', 'own_mass',
                         'customs_clearance_date', 'country_of_origin', 'cn', 'procedure_code', 'requested_procedure',
                         'previous_procedure', 'vat_base', 'vat_amount', 'currency_vat', 'customs_duty_base',
                         'customs_duty_amount', 'currency_duty', 'position_system_id', 'material_number']



class MICParser:
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        self.worksheet = self._get_mic_sheet()

    def parse(self) -> list:
        mic_data = []
        file_header = self._check_mic_header()
        file_header = self._normalize_mic_header(file_header)
        for row in list(self.worksheet.rows)[1:]:
            mic_row = dict()

            for i in range(len(file_header)):
                field = file_header[i]
                value = row[i].value
                mic_row[field] = value

            if self._valid_row(mic_row):
                mic_data.append(mic_row)
        return mic_data


    def _get_mic_sheet(self):
        if len(self.workbook.sheetnames) != 1:
            raise ParserException(self.file_path, 'The file must contain one sheet only.')

        return self.workbook[self.workbook.sheetnames[0]]

    def _check_mic_header(self):
        file_header = []

        for field in list(self.worksheet.rows)[0]:
            field_value = field.value
            file_header.append(field_value)

        lacking_fields = set(mic_original_header) - set(file_header)
        if len(lacking_fields) > 0:
            raise ParserException(self.file_path, f'The header does not contain the required fields. '
                                                  f'Lacking fields: {lacking_fields}')

        return file_header

    @staticmethod
    def _normalize_mic_header(mic_header: list) -> list:
        return normalize_list(mic_header)

    @staticmethod
    def _valid_row(mic_row: dict) -> bool:
        for value in mic_row.values():
            if value is not None:
                return True
        return False


if __name__ == '__main__':
    parser = MICParser(Path('/home/fabad/Descargas/MICFinal.xlsx'))
    mic_data = parser.parse()
    print(mic_data)