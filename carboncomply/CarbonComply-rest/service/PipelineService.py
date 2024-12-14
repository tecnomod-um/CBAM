import csv
import os
from pathlib import Path

import openpyxl
import rdflib
from rdflib import Graph
from xsdata.formats.dataclass.serializers.config import SerializerConfig

from CommunicationReport import FileNames
from CommunicationReport.FileNames import COMMUNICATION_REPORT_EXPORT_CSV_CONFIG, MIC_EXPORT_CSV_CONFIG
from CommunicationReport.enricher.CommunicationReportEnricher import CommunicationReportEnricher
from CommunicationReport.parser.CommunicationReportParser import CommunicationReportParser, \
    summary_communication_sheet_name, instance_data_sheet_name
from mic.enricher.MICEnricher import MICEnricher
from mic.parser.MICParser import MICParser
from service.CSV2RDFMappingService import CSV2RDFMappingService
from service.FileService import FileService
from service.RDF2XMLMappingService import RDF2XMLMappingService
from utils.Singleton import Singleton
from xsdata.formats.dataclass.serializers import XmlSerializer


# TODO: Include more files when available!
mappings_per_file = {
    FileNames.CBAM_SUMMARY_OF_PRODUCTS: [Path(Path(__file__).parent, '../resources/mapping_definition/summary_of_products.yaml')],
    # FileNames.CBAM_INSTALLATION_DETAILS: [Path(Path(__file__).parent, '../resources/mapping_definition/installation_details.yaml')],
    FileNames.CBAM_SUMMARY_OF_EMISSIONS: [Path(Path(__file__).parent, '../resources/mapping_definition/summary_of_emissions.yaml')],
    FileNames.CBAM_ABOUT_THE_INSTALLATION: [Path(Path(__file__).parent, '../resources/mapping_definition/about_the_installation.yaml')],
    FileNames.CBAM_REPORTING_PERIOD: [Path(Path(__file__).parent, '../resources/mapping_definition/reporting_period.yaml')],
    FileNames.MIC: [Path(Path(__file__).parent, '../resources/mapping_definition/mic.yaml')]
}


class PipelineService(metaclass=Singleton):

    def execute_pipeline(self, workdir, files: list) -> dict:
        files_by_type = self._separate_files_by_type(files)
        local_folders = self._create_local_folders(workdir, files_by_type)
        self._create_csvs(workdir, local_folders)
        rdf = self._generate_rdf(workdir, local_folders)
        xml = self._generate_xml(workdir)
        return {
            'rdf': rdf,
            'xml': xml
        }


    def _separate_files_by_type(self, files: list) -> dict:
        files_by_type = {
            'mic': [],
            'cbam': []
        }
        for file in files:
            excel_workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
            if summary_communication_sheet_name in excel_workbook.sheetnames and instance_data_sheet_name in excel_workbook.sheetnames:
                files_by_type['cbam'].append(file)
            else:
                files_by_type['mic'].append(file)
            excel_workbook.close()
            file.seek(0) # Reset file read position... needed to continue using the file outside.

        return files_by_type


    def _create_local_folders(self, workdir:Path, files: dict) -> dict:
        local_folders = {
            'mic': self._create_mic_local_folders(workdir, files['mic']),
            'cbam': self._create_cbam_local_folders(workdir, files['cbam'])
        }
        return local_folders


    def _create_cbam_local_folders(self, workdir:Path, files: list):
        file_service = FileService()
        created_folders = []
        for i in range(len(files)):
            file = files[i]
            folder = file_service.create_folder(workdir, f'summary_communications_{str(i)}')
            created_folders.append(folder)
            file_service.save(folder, file, FileNames.CBAM_COMMUNICATION_XLSX)
        return created_folders

    def _create_mic_local_folders(self, workdir:Path, files: list):
        file_service = FileService()
        created_folders = []
        for i in range(len(files)):
            file = files[i]
            folder = file_service.create_folder(workdir, f'mic_{str(i)}')
            created_folders.append(folder)
            file_service.save(folder, file, FileNames.MIC_XLSX)
        return created_folders


    def _create_csvs(self, workdir, local_folders: dict):
        self._create_cbam_comm_report_csvs(workdir, local_folders['cbam'])
        self._create_mic_csvs(workdir, local_folders['mic'])


    def _create_cbam_comm_report_csvs(self, workdir, cbam_local_folders):
        for cbam_local_folder in cbam_local_folders:
            cbam_comm_report_path = Path(cbam_local_folder, FileNames.CBAM_COMMUNICATION_XLSX)
            parser = CommunicationReportParser(cbam_comm_report_path)
            cbam_comm_report = parser.parse()

            enricher = CommunicationReportEnricher(cbam_comm_report)
            cbam_comm_report_enriched = enricher.enrich()

            for key, fileinfo in COMMUNICATION_REPORT_EXPORT_CSV_CONFIG.items():
                with open(Path(cbam_local_folder, fileinfo['filename']), 'w', newline='') as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=fileinfo['header'], delimiter=",")
                    writer.writeheader()
                    writer.writerows(cbam_comm_report_enriched[key])


    def _create_mic_csvs(self, workdir, mic_local_folders):
        for mic_local_folder in mic_local_folders:
            mic_path = Path(mic_local_folder, FileNames.MIC_XLSX)
            parser = MICParser(mic_path)
            mic = parser.parse()
            enricher = MICEnricher(mic)
            mic_enriched = enricher.enrich()

            with open(Path(mic_local_folder, MIC_EXPORT_CSV_CONFIG['filename']), 'w', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=MIC_EXPORT_CSV_CONFIG['header'], delimiter=",")
                writer.writeheader()
                writer.writerows(mic_enriched)



    def _generate_rdf(self, workdir: Path, local_folders_per_file_type: dict) -> str:
        local_folders = local_folders_per_file_type['cbam'] + local_folders_per_file_type['mic']
        mapping_service = CSV2RDFMappingService()
        file_service = FileService()
        graph: Graph = Graph()
        for local_folder in local_folders:
            for filename, mapping_files in mappings_per_file.items():
                if not Path(local_folder, filename).exists():
                    continue
                local_graph: Graph = mapping_service.convert(local_folder, filename, mapping_files)
                local_graph_filepath = Path(local_folder, filename + '.rdf')
                local_graph.serialize(destination=local_graph_filepath, format='turtle', encoding='utf-8')
                graph.parse(local_graph_filepath, format='turtle', encoding='utf-8')
                local_graph_filepath.unlink()

        rdf_output_path = Path(workdir, FileNames.CBAM_COMMUNICATION_RDF)
        self._define_prefixes(graph)
        graph.serialize(destination=rdf_output_path, format='turtle', encoding='utf-8')

        return file_service.read_file(rdf_output_path)

    def _generate_xml(self, workdir: Path) -> str:
        file_service = FileService()
        mapping_service = RDF2XMLMappingService()
        xml_output_path = Path(workdir, FileNames.CBAM_COMMUNICATION_XML)
        mapping_file = Path(Path(__file__).parent, '../resources/mapping_definition/rdf2xml.xs')
        xml_model = mapping_service.convert(workdir, FileNames.CBAM_COMMUNICATION_RDF, mapping_file)
        serializer = XmlSerializer(config=SerializerConfig(pretty_print=True))
        with open(xml_output_path, 'w+') as xml_output_file:
            serializer.write(xml_output_file, xml_model)

        return file_service.read_file(xml_output_path)

    def _define_prefixes(self, graph: Graph):
        schema = rdflib.Namespace('http://schema.org/')
        geonames_ont = rdflib.Namespace('https://www.geonames.org/ontology#')
        geonames = rdflib.Namespace('https://sws.geonames.org/')
        ontology_se_cbam = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/')
        ontology_se_cbam_emissionsqualifyingparameters = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/emissionsqualifyingparameters/')
        ontology_se_cbam_cn = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/CN/')
        ontology_se_cbam_cbamreport = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/cbamreport/')
        ontology_se_cbam_electricity_determination = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/Electricity_Determination/')
        ontology_se_cbam_instrument = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/Instrument/')
        ontology_se_cbam_country = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/CountryPlus/')
        ontology_se_cbam_currency = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/currency/')
        ontology_se_cbam_production_method = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/Production_Method/')
        ontology_se_cbam_cbam_goods = rdflib.Namespace('https://ontology.siemens-energy.com/cbam/cbam_goods/')
        se_cbam = rdflib.Namespace('https://data.siemens-energy.com/cbam#')
        se_cbam_emissionsqualifyingparameters = rdflib.Namespace('https://data.siemens-energy.com/cbam/emissionsqualifyingparameters#')
        se_cbam_cn = rdflib.Namespace('https://data.siemens-energy.com/cbam/CN#')
        se_cbam_cbamreport = rdflib.Namespace('https://data.siemens-energy.com/cbam/cbamreport#')
        se_cbam_electricity_determination = rdflib.Namespace('https://data.siemens-energy.com/cbam/Electricity_Determination#')
        se_cbam_instrument = rdflib.Namespace('https://data.siemens-energy.com/cbam/Instrument#')
        se_cbam_country = rdflib.Namespace('https://data.siemens-energy.com/cbam/CountryPlus#')
        se_cbam_currency = rdflib.Namespace('https://data.siemens-energy.com/cbam/currency#')
        se_cbam_production_method = rdflib.Namespace('https://data.siemens-energy.com/cbam/Production_Method#')
        se_cbam_cbam_goods = rdflib.Namespace('https://data.siemens-energy.com/cbam/cbam_goods#')
        graph.bind('sch', schema)
        graph.bind('ontology-geo', geonames_ont)
        graph.bind('geo', geonames, replace=True)
        graph.bind('ontology-se-cbam', ontology_se_cbam)
        graph.bind('ontology-se-cbam_emissionsqualifyingparameters', ontology_se_cbam_emissionsqualifyingparameters)
        graph.bind('ontology-se-cbam-cn', ontology_se_cbam_cn)
        graph.bind('ontology-se-cbam-cbamreport', ontology_se_cbam_cbamreport)
        graph.bind('ontology-se-cbam-electricity-determination', ontology_se_cbam_electricity_determination)
        graph.bind('ontology-se-cbam-instrument', ontology_se_cbam_instrument)
        graph.bind('ontology-se-cbam-country', ontology_se_cbam_country)
        graph.bind('ontology-se-cbam-currency', ontology_se_cbam_currency)
        graph.bind('ontology-se-cbam-production-method', ontology_se_cbam_production_method)
        graph.bind('ontology-se-cbam-cbam-goods', ontology_se_cbam_cbam_goods)
        graph.bind('se-cbam', se_cbam)
        graph.bind('se-cbam-emissionsqualifyingparameters', se_cbam_emissionsqualifyingparameters)
        graph.bind('se-cbam-cn', se_cbam_cn)
        graph.bind('se-cbam-cbamreport', se_cbam_cbamreport)
        graph.bind('se-cbam-electricity-determination', se_cbam_electricity_determination)
        graph.bind('se-cbam-instrument', se_cbam_instrument)
        graph.bind('se-cbam-country', se_cbam_country)
        graph.bind('se-cbam-currency', se_cbam_currency)
        graph.bind('se-cbam-production-method', se_cbam_production_method)
        graph.bind('se-cbam-cbam-goods', se_cbam_cbam_goods)


if __name__ == '__main__':
    service = PipelineService()
    service.execute_pipeline(Path('/home/fabad/Descargas/CC/francisco.abad@um.es'))
